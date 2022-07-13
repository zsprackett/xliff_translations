#!/usr/bin/env python3

import openpyxl
import xml.etree.cElementTree as ET
import pprint
import re
import sys
import os
import getopt
import yaml

def get_lang(s):
    m = re.search('\s*([^\[]+)\s+\[\s*([^\[\]\s]+)\s*', s)
    if m:
        return { m[1].strip(): m[2].strip() }
    else:
        return {}
    
input_xliff = None
input_xlsx = None

try:
    opts, args = getopt.getopt(sys.argv[1:], "he:x:", ["english=", "excel="])
except getopt.GetoptError:
    print(f'{sys.argv[0]} -e english.xliff -x excel.xlsx')
    sys.exit(2)

for opt, arg in opts:
    if opt == '-h':
        sys.exit()
    elif opt in ("-e", "--english"):
        input_xliff = arg
    elif opt in ("-x", "--excel"):
        input_xlsx = arg

if input_xliff == None or input_xlsx == None:
    print(f'Filename arguments --english and --excel are required.')
    sys.exit(2)
if not os.path.isfile(input_xliff):
    print(f'Not a file: {input_xliff}')
    sys.exit(2)

langs = {}
with open(r'languages.yaml') as file:
    data = yaml.load(file, Loader=yaml.FullLoader)
    langs = data['languages']

workbook = openpyxl.load_workbook(input_xlsx)
worksheet = workbook.active

pp = pprint.PrettyPrinter(indent=4)

eng_strings = {}
eng_cell = None
for col in worksheet.iter_cols(1, worksheet.max_column):
    l = get_lang(col[0].value)
    if (l and list(l.values())[0] == 'en-US'):
        eng_cell = col[0]
        break

# skip our header
if eng_cell:
    for x in range(2,worksheet.max_row + 1):
        eng_cell = worksheet.cell(row=x, column=eng_cell.column)
        eng_strings[eng_cell.value] = eng_cell.row

if not len(eng_strings.keys()):
    print("Failed to find english strings!")
    sys.exit(1)

# Get the namespace, rather than hardcode it to "urn:oasis:names:tc:xliff:document:1.2"
tree = ET.parse(input_xliff)
root=tree.getroot()
namespace = re.match(r'{(.*)}', root.tag).group(1)
# Now parse again with the namespace set
ET.register_namespace("", namespace)
namespace = '{' + f"{namespace}" + '}'
tree = ET.parse(input_xliff)
tree._setroot(root)

for col in worksheet.iter_cols(eng_cell.column + 1, worksheet.max_column):
    lang_cell = col[0]
    l = get_lang(lang_cell.value)

    for file in root.findall(f"{namespace}file"):
        targetlang = file.attrib['target-language']
        file.set('target-language', list(l.values())[0])
        for body in file.iter(f"{namespace}body"):
            for group in body.iter(f"{namespace}group"):
                for transunit in body.iter(f"{namespace}trans-unit"):
                    source = transunit.find(f"{namespace}source")
                    target = transunit.find(f"{namespace}target")
                    try:
                         row = eng_strings[source.text]
                         ls = worksheet.cell(row=row, column=lang_cell.column)
                         target.text = ls.value
                    except:
                         pass

                    if not target.text or not target.text.strip():
                         target.text = source.text


    tree.write(list(l.values())[0] + ".xliff")
