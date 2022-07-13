#!/usr/bin/env python3

import openpyxl
import xml.etree.cElementTree as ET
import pprint
import re
import getopt
import os
import sys

def alpha_dict(list_arg):
    alphabet = []
    for i in range(0, len(list_arg)):
        alphabet.append(openpyxl.utils.cell.get_column_letter(i + 1))
    return dict(zip(alphabet, list_arg))

input_xliff = None
output_xlsx = None

try:
    opts, args = getopt.getopt(sys.argv[1:], "he:x:", ["english=", "excel="])
except getopt.GetoptError:
    print(f'{sys.argv[0]} -e english.xliff -x excel.xlsx')
    sys.exit(2)

for opt, arg in opts:
    if opt == '-h':
        sys.exit()
    elif opt in ("-e", "--english"):
        input_xliff = arg
    elif opt in ("-x", "--excel"):
        output_xlsx = arg

if input_xliff == None or output_xlsx == None:
    print(f'ERROR: Filename arguments --english and --excel are required.')
    sys.exit(2)
if not os.path.isfile(input_xliff):
    print(f'ERROR: Not a file: {input_xliff}')
    sys.exit(2)
if os.path.isfile(output_xlsx):
    print(f'ERROR: Please delete existing output file {output_xlsx}')
    sys.exit(2)

# Get the namespace, rather than hardcode it to "urn:oasis:names:tc:xliff:document:1.2"
tree = ET.parse('translated.xliff')
root=tree.getroot()
namespace = re.match(r'{(.*)}', root.tag).group(1)
# Now parse again with the namespace set
ET.register_namespace("", namespace)
namespace = '{' + f"{namespace}" + '}'
tree = ET.parse('translated.xliff')

workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "Strings"

not_translatable = [
    'insert.io.ImageView',
    'ButtonLink'
]

row = 1
fieldnames = alpha_dict ([
    'File',
    'Context',
    'English',
    'French',
    'German',
    'Portuguese',
    'Spanish'
])
for key in fieldnames:
    (col, row) = openpyxl.utils.cell.coordinate_from_string(key + str(1))
    col = openpyxl.utils.cell.column_index_from_string(col)
    cell = worksheet.cell(row=row, column=col)
    cell.fill = openpyxl.styles.PatternFill("solid", fgColor="000000")
    cell.font = openpyxl.styles.Font(b=True, color="FFFFFF")
    cell.value = fieldnames[key]

row = row + 1
for file in root.findall(f"{namespace}file"):
    sourcelang = file.attrib['source-language']
    for body in file.iter(f"{namespace}body"):
        for group in body.iter(f"{namespace}group"):
            for transunit in body.iter(f"{namespace}trans-unit"):
                source = transunit.find(f"{namespace}source")
                note = transunit.find(f"{namespace}note")
                if (not note.text in not_translatable):
                    cell = worksheet.cell(row=row, column=2)
                    cell.value = note.text
                    cell = worksheet.cell(row=row, column=3)
                    cell.value = source.text
                    row = row + 1

workbook.save(output_xlsx)
